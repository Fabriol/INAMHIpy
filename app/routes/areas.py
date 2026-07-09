from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, jsonify
from flask_login import login_required, current_user

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import send_file, current_app

# Modelos oficiales del sistema institucional
from app.models.base import db, SolicitudPazSalvo, Respuesta, LogAuditoria, Usuario, Rol

areas_bp = Blueprint('areas', __name__)

# ====================================================================
# 1. RUTA: PANEL INTELIGENTE DE TRÁMITES (CORREGIDO SIN BUSQUEDA EN_PROGRESO)
# ====================================================================
@areas_bp.route('/areas/tareas')
@login_required
def mis_tareas():
    # Lista de roles permitidos en el Centro de Gestión Institucional
    roles_areas = ['Administrativa', 'Financiera', 'TICs', 'Seguridad', 'Administrador', 'Talento Humano - Recepción Documentos', 'Ex Funcionario']
    
    if current_user.rol.nombre not in roles_areas:
        flash('Acceso denegado. No perteneces a un área de validación.', 'danger')
        return redirect(url_for('dashboard.index'))

    # FILTRADO DE BANDEJA SEGÚN EL ROL DE INICIO DE SESIÓN
    if current_user.rol.nombre in ['Administrador', 'Talento Humano - Recepción Documentos']:
        # El Administrador y RRHH visualizan la bandeja completa de la institución
        solicitudes = SolicitudPazSalvo.query.order_by(SolicitudPazSalvo.fecha_creacion.desc()).all()
        
    elif current_user.rol.nombre == 'Ex Funcionario':
        # El Ex Funcionario visualiza EXCLUSIVAMENTE su propia solicitud de Paz y Salvo
        solicitudes = SolicitudPazSalvo.query.filter_by(ex_funcionario_id=current_user.id).all()
        
    else:
        # Las áreas técnicas (TICs, Financiera, Seguridad, etc.) ven solo donde el Admin les delegó campos
        asignaciones = Respuesta.query.filter_by(usuario_asignado_id=current_user.id).all()
        
        # Obtenemos los IDs de los trámites asignados eliminando duplicados
        solicitud_ids = list(set([r.solicitud_id for r in asignaciones]))
        
        if solicitud_ids:
            solicitudes = SolicitudPazSalvo.query.filter(SolicitudPazSalvo.id.in_(solicitud_ids)).order_by(SolicitudPazSalvo.fecha_creacion.desc()).all()
        else:
            solicitudes = []
        
    # Inyección forzada de datos de identidad para limpiar el "Dato No Vinculado" en las tablas
    for sol in solicitudes:
        sol.usuario_data = Usuario.query.get(sol.ex_funcionario_id)
        
    return render_template('areas/pendientes.html', solicitudes=solicitudes)


# ====================================================================
# 2. RUTA: VISTA PREVIA AISLADA DEL DOCUMENTO (HOJA ESPEJO)
# ====================================================================
@areas_bp.route('/areas/ver/<int:solicitud_id>')
@login_required
def vista_previa(solicitud_id):
    solicitud = SolicitudPazSalvo.query.get_or_404(solicitud_id)
    solicitud.ex_funcionario = Usuario.query.get(solicitud.ex_funcionario_id)
    
    # Extraemos las respuestas asentadas para la previsualización de la Hoja Espejo
    respuestas_db = Respuesta.query.filter_by(solicitud_id=solicitud.id).all()
    datos_combinados = {r.campo_formulario: r.valor_respuesta for r in respuestas_db}
    
    return render_template('paz_salvo/ver_espejo.html', solicitud=solicitud, datos=datos_combinados)


# ====================================================================
# 3. RUTA: EMITIR DICTAMEN FINAL DE CIERRE (ADMINISTRADOR / RRHH)
# ====================================================================
@areas_bp.route('/areas/responder/<int:solicitud_id>', methods=['GET', 'POST'])
@login_required
def responder_preguntas(solicitud_id):
    if current_user.rol.nombre not in ['Administrador', 'Talento Humano - Recepción Documentos']:
        flash('No tiene permisos para emitir el dictamen final de este expediente.', 'danger')
        return redirect(url_for('areas.mis_tareas'))

    solicitud = SolicitudPazSalvo.query.get_or_404(solicitud_id)
    solicitud.ex_funcionario = Usuario.query.get(solicitud.ex_funcionario_id)

    # LA MAGIA: Extraemos las respuestas de la base de datos para pasarlas al HTML y calcular el porcentaje
    respuestas_db = Respuesta.query.filter_by(solicitud_id=solicitud.id).all()

    if request.method == 'POST':
        estado_final = request.form.get('estado_final')
        observacion_final = request.form.get('observacion_final')

        if not estado_final:
            flash('Error: Debe seleccionar un veredicto oficial para el trámite.', 'danger')
            return redirect(url_for('areas.responder_preguntas', solicitud_id=solicitud.id))

        solicitud.estado = str(estado_final).upper()
        detalle_auditoria = f"Emitió veredicto final: {estado_final}."

        if estado_final == 'Negado':
            solicitud.observacion_rechazo = str(observacion_final).upper()
            detalle_auditoria += f" Motivo de Rechazo: {observacion_final}"
            flash('Trámite negado de forma oficial. Se registró el motivo del rechazo en el sistema.', 'warning')
        elif estado_final == 'Aprobado':
            ex_funcionario = Usuario.query.get(solicitud.ex_funcionario_id)
            if ex_funcionario:
                ex_funcionario.activo = False
            flash('Trámite Aprobado Exitosamente. El expediente se ha cerrado y las credenciales del ex-funcionario fueron inhabilitadas.', 'success')

        from datetime import datetime, timedelta, timezone 
        ecuador_tz = timezone(timedelta(hours=-5))
        hora_real_ec = datetime.now(ecuador_tz).strftime("%d/%m/%Y %H:%M:%S")

        log = LogAuditoria(
            usuario_id=current_user.id, 
            modulo='Validación de Áreas', 
            accion='DICTAMEN FINAL', 
            detalle=f"El usuario {current_user.rol.nombre} procesó el trámite #{solicitud.id} a las {hora_real_ec}. {detalle_auditoria}"
        )
        db.session.add(log)
        db.session.commit()
        
        return redirect(url_for('areas.mis_tareas'))

    # ENVIAMOS LAS RESPUESTAS AL HTML PARA LA MATEMÁTICA
    return render_template('areas/responder.html', solicitud=solicitud, respuestas_db=respuestas_db)

# ====================================================================
# 4. RUTA: EXPORTAR BANDEJA DE TRÁMITES A EXCEL (SOLO ADMIN Y RRHH)
# ====================================================================
@areas_bp.route('/areas/exportar_excel')
@login_required
def exportar_tramites_excel():
    if current_user.rol.nombre not in ['Administrador', 'Talento Humano - Recepción Documentos']:
        flash('No tiene permisos para exportar la base de datos.', 'danger')
        return redirect(url_for('areas.mis_tareas'))

    solicitudes = SolicitudPazSalvo.query.order_by(SolicitudPazSalvo.fecha_creacion.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control de Trámites"

    fill_cabecera = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    font_cabecera = Font(name="Arial", size=11, color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    columnas = ["Nro. Trámite", "Cédula", "Nombres y Apellidos", "Correo Institucional", "Estado Actual", "Fecha de Creación"]
    ws.append(columnas)

    ws.row_dimensions[1].height = 25
    for celda in ws[1]:
        celda.fill = fill_cabecera
        celda.font = font_cabecera
        celda.alignment = align_center

    for sol in solicitudes:
        ex = Usuario.query.get(sol.ex_funcionario_id)
        ws.append([
            f"#{sol.id}",
            ex.cedula,
            f"{ex.nombres} {ex.apellidos}".upper(),
            ex.email,
            sol.estado,
            sol.fecha_creacion.strftime('%d/%m/%Y') if sol.fecha_creacion else 'N/A'
        ])

    for col in ws.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 25

    dir_temp = os.path.join(current_app.root_path, 'static', 'temp')
    os.makedirs(dir_temp, exist_ok=True)
    ruta_excel = os.path.join(dir_temp, 'Base_Tramites_INAMHI.xlsx')
    wb.save(ruta_excel)

    return send_file(os.path.abspath(ruta_excel), as_attachment=True, download_name="Base_Tramites_INAMHI.xlsx")