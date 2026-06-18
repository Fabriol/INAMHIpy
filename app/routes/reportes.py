import os
from datetime import datetime
from flask import Blueprint, send_file, flash, redirect, url_for, render_template, request, abort
from flask_login import login_required, current_user
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Librerías para el motor de PDF (ReportLab)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Importación de modelos
from app.models.base import db, SolicitudPazSalvo, Usuario, Pregunta, Respuesta, LogAuditoria

reportes_bp = Blueprint('reportes', __name__)

# ==========================================
# FUNCIÓN AUXILIAR: FILTRADO DE AUDITORÍA
# ==========================================
def obtener_logs_filtrados():
    query = LogAuditoria.query.join(Usuario)
    
    busqueda = request.args.get('usuario', '').strip()
    fecha_inicio = request.args.get('fecha_inicio', '').strip()
    fecha_fin = request.args.get('fecha_fin', '').strip()

    if busqueda:
        query = query.filter((Usuario.nombres.ilike(f'%{busqueda}%')) | (Usuario.cedula.ilike(f'%{busqueda}%')))
    if fecha_inicio:
        try:
            fecha_ini_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            query = query.filter(LogAuditoria.fecha >= fecha_ini_obj)
        except ValueError:
            pass
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(f"{fecha_fin} 23:59:59", '%Y-%m-%d %H:%M:%S')
            query = query.filter(LogAuditoria.fecha <= fecha_fin_obj)
        except ValueError:
            pass

    return query.order_by(LogAuditoria.fecha.desc()).all()


# ==========================================
# 1. VISTA: PANTALLA DE AUDITORÍA
# ==========================================
@reportes_bp.route('/admin/auditoria')
@login_required
def vista_auditoria():
    if current_user.rol.nombre not in ['Administrador', 'Talento Humano - Recepción Documentos']:
        abort(403)
        
    logs = obtener_logs_filtrados()
    return render_template('admin/auditoria.html', logs=logs)


# ==========================================
# 2. ACCIÓN: EXPORTAR TODA LA INF. A EXCEL (ESTRUCTURA 7.1 Y 7.2)
# ==========================================
@reportes_bp.route('/admin/auditoria/exportar-excel')
@login_required
def exportar_auditoria_excel():
    if current_user.rol.nombre not in ['Administrador', 'Talento Humano - Recepción Documentos']:
        abort(403)

    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # PESTAÑA 1: CONTROL DE TRÁMITES Y CAMPOS DE CADA ÁREA
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Control Procesos"
    
    # Estilos Corporativos (Muted, Dark Blue Fill)
    fill_cabecera = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    font_cabecera = Font(name="Arial", size=11, color="FFFFFF", bold=True)
    font_datos = Font(name="Arial", size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    border_fino = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Construcción dinámica de columnas por cada área configurada
    preguntas = Pregunta.query.filter_by(activa=True).order_by(Pregunta.rol_asignado_id).all()
    
    columnas_ws1 = [
        "Nro. Trámite", "Cédula Funcionario", "Apellidos y Nombres", "Correo Personal"
    ]
    # Inyección de los campos de cada área de forma dinámica
    for p in preguntas:
        columnas_ws1.append(f"[{p.rol.nombre}] {p.enunciado}")
        
    columnas_ws1.extend(["Estado del Proceso", "Fecha Creación", "Fecha Finalización"])
    ws1.append(columnas_ws1)

    # Estilar cabecera de la pestaña 1
    ws1.row_dimensions[1].height = 28
    for col_idx, celda in enumerate(ws1[1], 1):
        celda.fill = fill_cabecera
        celda.font = font_cabecera
        celda.alignment = align_center

    # Poblado de datos de trámites
    solicitudes = SolicitudPazSalvo.query.all()
    for sol in solicitudes:
        ex = sol.ex_funcionario
        fila = [
            f"#{sol.id}",
            ex.cedula,
            f"{ex.apellidos} {ex.nombres}",
            ex.email
        ]
        
        # Mapeo dinámico de respuestas por celda/pregunta
        for p in preguntas:
            resp = Respuesta.query.filter_by(solicitud_id=sol.id, pregunta_id=p.id).first()
            fila.append(resp.valor_respuesta if resp else "PENDIENTE")
            
        fila.extend([
            sol.estado,
            sol.fecha_creacion.strftime('%Y-%m-%d %H:%M') if sol.fecha_creacion else 'N/A',
            sol.fecha_cierre.strftime('%Y-%m-%d %H:%M') if getattr(sol, 'fecha_cierre', None) else 'En proceso'
        ])
        ws1.append(fila)

    # Autoajuste de columnas pestaña 1
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
        for cell in col:
            if cell.row > 1:
                cell.font = font_datos
                cell.border = border_fino

    # ----------------------------------------------------
    # PESTAÑA 2: REGISTRO DE ACCIONES (LOGS)
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Registro de Acciones (Logs)")
    columnas_ws2 = ["ID Log", "Operario (Cédula)", "Rol Institucional", "Módulo afectado", "Acción Realizada", "Detalle / Modificaciones", "Fecha y Hora"]
    ws2.append(columnas_ws2)
    
    ws2.row_dimensions[1].height = 28
    for celda in ws2[1]:
        celda.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Charcoal para diferenciar
        celda.font = font_cabecera
        celda.alignment = align_center

    logs = obtener_logs_filtrados()
    for log in logs:
        ws2.append([
            f"#{log.id}",
            log.usuario.cedula,
            log.usuario.rol.nombre,
            log.modulo,
            log.accion,
            log.detalle,
            log.fecha.strftime('%Y-%m-%d %H:%M:%S')
        ])

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)
        for cell in col:
            if cell.row > 1:
                cell.font = font_datos
                cell.border = border_fino
                cell.alignment = align_left if cell.column == 6 else align_center

    # Guardado seguro
    dir_temp = os.path.join('app', 'static', 'temp')
    os.makedirs(dir_temp, exist_ok=True)
    ruta_excel = os.path.join(dir_temp, 'Auditoria_General_INAMHI.xlsx')
    wb.save(ruta_excel)

    return send_file(os.path.abspath(ruta_excel), as_attachment=True, download_name="Auditoria_General_INAMHI.xlsx")


# ==========================================
# 3. ACCIÓN: EXPORTAR AUDITORÍA A PDF
# ==========================================
@reportes_bp.route('/admin/auditoria/exportar-pdf')
@login_required
def exportar_auditoria_pdf():
    if current_user.rol.nombre not in ['Administrador', 'Talento Humano - Recepción Documentos']:
        abort(403)

    logs = obtener_logs_filtrados()
    
    dir_temp = os.path.join('app', 'static', 'temp')
    os.makedirs(dir_temp, exist_ok=True)
    ruta_pdf = os.path.join(dir_temp, 'Auditoria_Sistema_INAMHI.pdf')

    # Configuración de hoja horizontal (landscape) para albergar la grilla completa
    doc = SimpleDocTemplate(
        ruta_pdf, pagesize=landscape(A4),
        rightMargin=20, leftMargin=20, topMargin=25, bottomMargin=25
    )
    elementos = []
    estilos = getSampleStyleSheet()

    # Estilos de texto controlados
    style_title = ParagraphStyle(
        'TitleStyle', parent=estilos['Heading1'],
        fontName='Helvetica-Bold', fontSize=16, leading=20,
        textColor=colors.HexColor('#1E3A8A'), alignment=1, spaceAfter=6
    )
    style_subtitle = ParagraphStyle(
        'SubTitleStyle', parent=estilos['Normal'],
        fontName='Helvetica', fontSize=10, leading=14,
        textColor=colors.HexColor('#475569'), alignment=1, spaceAfter=20
    )
    style_cell = ParagraphStyle(
        'CellText', parent=estilos['Normal'],
        fontName='Helvetica', fontSize=8, leading=11, alignment=1
    )
    style_cell_left = ParagraphStyle(
        'CellTextLeft', parent=estilos['Normal'],
        fontName='Helvetica', fontSize=8, leading=11, alignment=0
    )

    elementos.append(Paragraph("INSTITUTO NACIONAL DE METEOROLOGÍA E HIDROLOGÍA - INAMHI", style_title))
    elementos.append(Paragraph(f"Reporte del Historial de Auditoría y Logs del Sistema — Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M')}", style_subtitle))

    # Definición de cabecera de tabla
    datos_tabla = [[
        Paragraph("<b>ID</b>", style_cell),
        Paragraph("<b>OPERARIO</b>", style_cell),
        Paragraph("<b>ROL</b>", style_cell),
        Paragraph("<b>MÓDULO</b>", style_cell),
        Paragraph("<b>ACCIÓN</b>", style_cell),
        Paragraph("<b>DETALLE / MODIFICACIONES</b>", style_cell),
        Paragraph("<b>FECHA Y HORA</b>", style_cell)
    ]]

    # Llenado de filas transformadas a Paragraphs seguros contra desbordamiento
    for log in logs:
        datos_tabla.append([
            Paragraph(f"#{log.id}", style_cell),
            Paragraph(log.usuario.cedula, style_cell),
            Paragraph(log.usuario.rol.nombre, style_cell),
            Paragraph(log.modulo, style_cell),
            Paragraph(log.accion, style_cell),
            Paragraph(log.detalle, style_cell_left),
            Paragraph(log.fecha.strftime('%Y-%m-%d %H:%M:%S'), style_cell)
        ])

    # Grilla con anchos medidos al ancho total A4 horizontal (aprox 800 puntos libres)
    tabla = Table(datos_tabla, colWidths=[40, 80, 110, 80, 100, 270, 100])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))
    
    elementos.append(tabla)
    doc.build(elementos)

    return send_file(os.path.abspath(ruta_pdf), as_attachment=True, download_name="Auditoria_Sistema_INAMHI.pdf")